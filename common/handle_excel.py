"""
============================
Author:丁琴
Time: 15:56
E-mail:394597923@qq.com
Company:南京瓦丁科技限公司
============================
"""
import openpyxl
class Handle_Excel:
    def __init__(self,filename,sheetname):
        """

        :param filename: 文件名
        :param sheetname: 表单名
        """
        self.filename=filename
        self.sheetname=sheetname
    def read_excle(self):
        # 加载工作簿
        wb=openpyxl.load_workbook(self.filename)
        # 获取表单
        sh=wb[self.sheetname]
        # 获取所有行
        row_data=list(sh.rows)
        case_data=[]
        # 获取表头数据
        title=[]
        for i in row_data[0]:
            title.append(i.value)
        # 获取其他行数据
        for item in row_data[1:]:
            values=[]
            for i in item:
                values.append(i.value)
            case=dict(zip(title,values))
            case_data.append(case)
        return case_data
    # 写入数据
    def wirte_excel(self,row,column,value):
        wb=openpyxl.load_workbook(self.filename)
        sh=wb[self.sheetname]
        sh.cell(row=row,column=column,value=value)
        wb.save(self.filename)

if __name__=="__main__":
    excel=Handle_Excel(r"F:\py27\py_api_test\data\cases.xlsx","recharge")
    res=excel.read_excle()
    print(res)


